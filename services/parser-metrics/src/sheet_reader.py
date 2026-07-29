"""
sheet_reader.py - Reads .xlsx files and identifies worksheet structure.
Outputs a WorkbookProfile with sheets, headers, data areas, merged cells.
"""
import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def read_workbook(file_path: str, job_id: str) -> dict[str, Any]:
    """
    Read an Excel workbook and produce a WorkbookProfile dict.
    
    Args:
        file_path: Path to the .xlsx file
        job_id: Job identifier
    
    Returns:
        WorkbookProfile dict matching the contract schema
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {file_path}")
    
    if not path.suffix.lower() in ('.xlsx', '.xlsm'):
        raise ValueError(f"Unsupported file format: {path.suffix}. Only .xlsx/.xlsm supported.")

    # Compute file hash
    file_hash = hashlib.sha256(path.read_bytes()).hexdigest()

    wb = load_workbook(file_path, data_only=True, read_only=False)
    
    sheets = []
    all_periods: set[str] = set()
    all_entities: set[str] = set()
    detected_units: dict[str, str] = {}

    for sheet_name in wb.sheetnames:
        ws: Worksheet = wb[sheet_name]
        sheet_profile = _analyze_sheet(ws, sheet_name)
        sheets.append(sheet_profile)
        
        # Collect periods from column headers (format: 5-digit like 11401)
        for col in sheet_profile["columns"]:
            if _is_period(col):
                all_periods.add(col)
        
        # Collect entities from first column data
        for row in ws.iter_rows(
            min_row=sheet_profile["dataStartRow"],
            max_row=sheet_profile["dataEndRow"],
            min_col=1,
            max_col=1,
            values_only=True
        ):
            if row[0] and isinstance(row[0], str) and not _is_period(str(row[0])):
                all_entities.add(str(row[0]).strip())

    wb.close()

    return {
        "profileId": f"profile-{job_id}",
        "jobId": job_id,
        "sourceFileUri": str(file_path),
        "sourceFileHash": file_hash,
        "sheets": sheets,
        "detectedPeriods": sorted(all_periods),
        "detectedEntities": sorted(all_entities),
        "detectedUnits": detected_units,
    }


def _analyze_sheet(ws: Worksheet, sheet_name: str) -> dict[str, Any]:
    """Analyze a single worksheet structure."""
    max_row = ws.max_row or 1
    max_col = ws.max_column or 1

    # Detect header row (first non-empty row)
    header_row = 1
    columns: list[str] = []
    
    for row_idx in range(1, min(max_row + 1, 10)):
        row_values = []
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                row_values.append(str(cell.value).strip())
        
        if len(row_values) >= 2:  # Found header row
            header_row = row_idx
            # Get all column headers
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                columns.append(str(cell.value).strip() if cell.value else "")
            break

    data_start_row = header_row + 1
    data_end_row = max_row

    # Find actual data end (skip empty trailing rows)
    for row_idx in range(max_row, data_start_row - 1, -1):
        has_data = False
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=row_idx, column=col_idx).value is not None:
                has_data = True
                break
        if has_data:
            data_end_row = row_idx
            break

    # Detect merged cells
    merged = [str(mc) for mc in ws.merged_cells.ranges]

    # Count nulls in data area
    null_count = 0
    for row_idx in range(data_start_row, data_end_row + 1):
        for col_idx in range(1, max_col + 1):
            if ws.cell(row=row_idx, column=col_idx).value is None:
                null_count += 1

    return {
        "sheetName": sheet_name,
        "headerRow": header_row,
        "dataStartRow": data_start_row,
        "dataEndRow": data_end_row,
        "columns": columns,
        "mergedCells": merged,
        "dataQuality": {
            "nullCount": null_count,
            "formatIssues": [],
        },
    }


def _is_period(value: str) -> bool:
    """Check if a string looks like a ROC period (5 digits like 11401)."""
    if len(value) == 5 and value.isdigit():
        year = int(value[:3])
        month = int(value[3:])
        return 100 <= year <= 200 and 1 <= month <= 12
    return False
