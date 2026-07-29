"""
source_mapper.py - Creates SourceRef records for every normalized value.
Maps each value back to its original sheet, cell address, and raw content.
"""
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from .normalizer import normalize_number, normalize_period, normalize_entity, detect_unit


def build_source_refs(
    file_path: str,
    profile: dict[str, Any],
) -> list[dict[str, Any]]:
    """
    Read the workbook and build SourceRef for every data cell.
    
    Args:
        file_path: Path to the .xlsx file
        profile: WorkbookProfile dict from sheet_reader
    
    Returns:
        List of SourceRef dicts
    """
    wb = load_workbook(file_path, data_only=True, read_only=False)
    source_refs: list[dict[str, Any]] = []
    ref_counter = 0

    for sheet_info in profile["sheets"]:
        sheet_name = sheet_info["sheetName"]
        ws = wb[sheet_name]
        columns = sheet_info["columns"]
        header_row = sheet_info["headerRow"]
        data_start = sheet_info["dataStartRow"]
        data_end = sheet_info["dataEndRow"]

        # Determine which columns are period columns vs entity/label columns
        period_cols: dict[int, str] = {}  # col_index -> period string
        entity_col: Optional[int] = None

        for col_idx, col_name in enumerate(columns, start=1):
            period = normalize_period(col_name)
            if period:
                period_cols[col_idx] = period
            elif col_idx == 1:
                entity_col = col_idx

        # Iterate data rows
        for row_idx in range(data_start, data_end + 1):
            # Get entity from first column
            entity_raw = ws.cell(row=row_idx, column=1).value
            if entity_raw is None:
                continue
            entity = normalize_entity(str(entity_raw).strip())

            # Process each period column
            for col_idx, period in period_cols.items():
                cell = ws.cell(row=row_idx, column=col_idx)
                raw_value = cell.value
                
                if raw_value is None:
                    continue

                normalized_value, data_type = normalize_number(raw_value)
                if normalized_value is None:
                    continue

                ref_counter += 1
                cell_address = f"{get_column_letter(col_idx)}{row_idx}"

                source_refs.append({
                    "sourceId": f"src-{ref_counter:06d}",
                    "sheetName": sheet_name,
                    "cellAddress": cell_address,
                    "rawValue": str(raw_value),
                    "normalizedValue": normalized_value,
                    "dataType": data_type,
                    "period": period,
                    "entity": entity,
                })

    wb.close()
    return source_refs
