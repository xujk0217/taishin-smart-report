from __future__ import annotations

import pytest
from openpyxl import Workbook

from lobster_runtime.calculation import CalculationOutput, execute_program, validate_program


PROGRAM = '''
def calculate(workbooks):
    workbook = openpyxl.load_workbook(workbooks["upload-1"], read_only=True, data_only=True, keep_links=False)
    worksheet = workbook["Data"]
    source_rows = list(worksheet.values)
    total = 0
    for row in source_rows[1:]:
        total += float(row[1] or 0)
    return {"tasks": [{"task_id": "task-total", "metric_id": "m-total", "formula_id": "f-total", "rows": [{"value": total, "unit": "count"}], "warnings": []}]}
'''


def test_validated_program_reads_only_the_supplied_workbook_and_returns_json(tmp_path):
    source = tmp_path / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Data"
    worksheet.append(["name", "value"])
    worksheet.append(["A", 2])
    worksheet.append(["B", 3])
    workbook.save(source)

    validate_program(PROGRAM)
    result = execute_program(PROGRAM, {"upload-1": source})

    assert result.tasks[0].rows == [{"value": 5.0, "unit": "count"}]


def test_program_with_import_or_file_write_is_rejected():
    validate_program("import openpyxl\ndef calculate(workbooks):\n openpyxl.load_workbook(workbooks['x'], read_only=True)\n return {'tasks': []}")
    with pytest.raises(ValueError, match="CALCULATION_CODE_UNSAFE_IMPORT"):
        validate_program("import os\ndef calculate(workbooks):\n return {'tasks': []}")
    with pytest.raises(ValueError, match="CALCULATION_CODE_UNSAFE_ATTRIBUTE"):
        validate_program("def calculate(workbooks):\n openpyxl.save('x')\n return {'tasks': []}")


def test_program_can_use_safe_openpyxl_from_import_and_context_manager():
    validate_program(
        "from openpyxl import load_workbook\n"
        "def calculate(workbooks):\n"
        " with load_workbook(workbooks['x'], read_only=True) as workbook:\n"
        "  return {'tasks': []}\n"
    )
    with pytest.raises(ValueError, match="CALCULATION_CODE_UNSAFE_IMPORT"):
        validate_program("from os import getenv\ndef calculate(workbooks):\n return {'tasks': []}")


def test_program_can_use_lambda_as_a_local_sort_key():
    validate_program(
        "import openpyxl\n"
        "def calculate(workbooks):\n"
        " rows = sorted([], key=lambda row: row.get('value', 0))\n"
        " openpyxl.load_workbook(workbooks['x'], read_only=True)\n"
        " return {'tasks': []}\n"
    )


def test_program_with_unbounded_loop_is_rejected_before_execution():
    with pytest.raises(ValueError, match="CALCULATION_CODE_UNSAFE_SYNTAX:While"):
        validate_program(
            "import openpyxl\n"
            "def calculate(workbooks):\n"
            " while True:\n"
            "  pass\n"
            " openpyxl.load_workbook(workbooks['x'], read_only=True)\n"
            " return {'tasks': []}\n"
        )


def test_result_can_keep_actionable_data_quality_warnings():
    output = CalculationOutput.model_validate({
        "tasks": [{
            "task_id": "task-1", "metric_id": "metric-1", "formula_id": "formula-1",
            "rows": [], "warnings": [f"warning-{index}" for index in range(32)],
        }],
    })
    assert len(output.tasks[0].warnings) == 32


def test_execution_exposes_only_a_missing_identifier_for_name_errors(tmp_path):
    source = tmp_path / "source.xlsx"
    Workbook().save(source)
    program = '''
def calculate(workbooks):
    openpyxl.load_workbook(workbooks["upload-1"], read_only=True)
    return missing_helper
'''
    with pytest.raises(RuntimeError, match="CALCULATION_EXECUTION_FAILED:NameError:missing_helper"):
        execute_program(program, {"upload-1": source})
