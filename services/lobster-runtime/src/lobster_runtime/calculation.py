"""Generate, constrain, and execute calculation programs for an approved plan."""

from __future__ import annotations

import ast
import hashlib
import json
import multiprocessing
import time
from dataclasses import dataclass
from pathlib import Path
from queue import Empty
from typing import TYPE_CHECKING, Any
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, ValidationError, model_validator
if TYPE_CHECKING:
    from strands.models import Model

from .contracts import AIPlanningOutput

MAX_PROGRAM_CHARACTERS = 48_000
# Generation transport/structured-output failures and generated-program
# execution failures have independent retry budgets. The single generation
# retry is shared across the run so one execution correction cannot expand the
# worst case beyond three model calls and two bounded child executions.
MAX_GENERATION_RETRIES = 1
MAX_EXECUTION_CORRECTIONS = 1
# Large workbooks can contain tens of thousands of rows. The child remains
# process-isolated and is forcibly terminated at this bound.
MAX_EXECUTION_SECONDS = 120
MAX_RESULT_ROWS_PER_TASK = 12_000
MAX_PREVIEW_ROWS = 12

CALCULATION_CODE_PROMPT = """
You are the Lobster Calculation Code Agent. Generate one complete Python program for the supplied, already
approved calculation plan. Return CalculationProgram only.

The program must define calculate(workbooks), where workbooks maps upload IDs to local .xlsx paths. It must
return exactly {"tasks": [...]}; every task object must include task_id, metric_id, formula_id, rows, and
warnings. Each task must match the supplied task ID, metric ID, formula ID and output_fields. Read Excel only
with openpyxl.load_workbook(path, read_only=True, data_only=True, keep_links=False). You may use either
`import openpyxl` or `from openpyxl import load_workbook`; do not import any other module,
open files, use network access, execute commands, use reflection, or modify workbooks. Use only basic Python
and the injected openpyxl module. Compute values from workbooks, never embed source-data numbers.

HARD SYNTAX RULE: The generated source must not contain the `while` keyword anywhere. Use only `for` loops over
finite collections. If you would normally use `while`, rewrite it as a `for` loop before returning the program;
programs containing `while` are rejected before they can run.

FAST, COMPATIBLE PROGRAM RULES: Use only the injected basic built-ins and openpyxl; do not reference pandas,
numpy, math, statistics, collections, defaultdict, Counter, or any helper that has not been defined in your
source. Iterate only over workbook rows, worksheet columns, or lists built from them: never use a large fixed
numeric range to search for data. Read each sheet once, reuse parsed rows and aggregates across all tasks, and
close each workbook after reading. Cap warnings at 100 by aggregating repeated warnings rather than appending
one warning per row. A small workbook must finish quickly; prefer direct dictionaries and simple for-loops.

Respect every input binding, formula expression, code-generation instruction, validation check, and provenance
requirement. For every result row, include only JSON scalar values. Put non-fatal data-quality findings in
warnings. Load each workbook and worksheet at most once, make a single pass to build reusable aggregates, and
derive all task outputs from those aggregates rather than repeatedly scanning rows. Keep the program concise
and deterministic. Return only executable source in the CalculationProgram payload: no prose, Markdown,
comments, docstrings, example data, or explanatory helper code. Define calculate(workbooks) plus only the
small helpers it directly calls. Target under 8,000 source characters; prefer shared aggregates over one
near-duplicate block per task.
""".strip()


class CalculationProgram(BaseModel):
    model_config = ConfigDict(extra="forbid")
    python_code: str = Field(min_length=80, max_length=MAX_PROGRAM_CHARACTERS)


class TaskResult(BaseModel):
    model_config = ConfigDict(extra="forbid")
    task_id: str = Field(min_length=1, max_length=160)
    metric_id: str = Field(min_length=1, max_length=160)
    formula_id: str = Field(min_length=1, max_length=160)
    rows: list[dict[str, Any]] = Field(max_length=MAX_RESULT_ROWS_PER_TASK)
    warnings: list[str] = Field(default_factory=list, max_length=100)


class CalculationOutput(BaseModel):
    model_config = ConfigDict(extra="forbid")
    tasks: list[TaskResult] = Field(max_length=30)


@dataclass(frozen=True)
class CalculationArtifact:
    execution_id: str
    generated_at: str
    duration_ms: int
    code_sha256: str
    python_code: str
    tasks: list[TaskResult]

    def artifact_payload(self) -> dict[str, Any]:
        return {
            "artifact_version": "calculation-artifact-v1",
            "execution_id": self.execution_id,
            "generated_at": self.generated_at,
            "duration_ms": self.duration_ms,
            "code_sha256": self.code_sha256,
            "python_code": self.python_code,
            "tasks": [task.model_dump(mode="json") for task in self.tasks],
        }

    def summary(self, artifact_key: str) -> dict[str, Any]:
        return {
            "executionId": self.execution_id,
            "status": "succeeded",
            "generatedAt": self.generated_at,
            "durationMs": self.duration_ms,
            "codeSha256": self.code_sha256,
            "codePreview": self.python_code[:8_000],
            "artifactKey": artifact_key,
            "tasks": [{
                "taskId": task.task_id,
                "metricId": task.metric_id,
                "formulaId": task.formula_id,
                "rowCount": len(task.rows),
                "preview": task.rows[:MAX_PREVIEW_ROWS],
                "warnings": task.warnings,
            } for task in self.tasks],
        }


def generate_and_execute(
    model: "Model",
    planning_output: AIPlanningOutput,
    workbook_paths: dict[str, Path],
    *,
    job_id: str | None = None,
    fallback_model: "Model | None" = None,
    previous_error: str | None = None,
) -> CalculationArtifact:
    """Generate and run calculation code with independent retry budgets."""
    context = {
        "calculation_plan": planning_output.calculation_plan.model_dump(mode="json"),
        "formula_plan": planning_output.formula_plan.model_dump(mode="json"),
        "metrics": [metric.model_dump(mode="json") for metric in planning_output.prompt_contract.metrics],
        "available_workbooks": {upload_id: str(path.name) for upload_id, path in workbook_paths.items()},
    }
    from strands import Agent
    started = time.monotonic()
    base_request = json.dumps(context, ensure_ascii=False, separators=(",", ":"))
    if previous_error:
        prior_reason = previous_error.replace("\n", " ")[:240]
        base_request += (
            "\nA previous manual calculation attempt failed with this deterministic error code: "
            f"{prior_reason}. Use it as correction feedback while preserving every approved task ID."
        )

    generation_attempt = 0
    generation_retries = 0
    execution_corrections = 0
    request = base_request

    while True:
        # A correction starts with the larger fallback model. Generation retry
        # has one independent budget shared across the complete run, so a
        # generation failure never consumes the execution-correction budget or
        # creates an unbounded model × execution retry product.
        primary_model = fallback_model if execution_corrections and fallback_model is not None else model
        alternate_model = model if primary_model is fallback_model else fallback_model
        attempt_models = [primary_model]
        if alternate_model is not None and generation_retries < MAX_GENERATION_RETRIES:
            attempt_models.append(alternate_model)
        generation_request = request
        program: CalculationProgram | None = None

        for model_index, attempt_model in enumerate(attempt_models):
            generation_attempt += 1
            _emit(job_id, "calculation-code", "started")
            try:
                agent = Agent(
                    model=attempt_model,
                    tools=[],
                    system_prompt=CALCULATION_CODE_PROMPT,
                    callback_handler=None,
                    load_tools_from_directory=False,
                    name="Lobster Calculation Code Agent",
                )
                result = agent(
                    generation_request,
                    structured_output_model=CalculationProgram,
                    idempotency_token=_sha256({
                        "job_id": job_id,
                        "calculation_plan": context["calculation_plan"],
                        "generation_attempt": generation_attempt,
                        "execution_correction": execution_corrections,
                    }),
                )
                if result.structured_output is None:
                    raise RuntimeError("CALCULATION_CODE_NOT_RETURNED")
                program = CalculationProgram.model_validate(result.structured_output)
                validate_program(program.python_code)
                _emit(job_id, "calculation-code", "completed")
                break
            except Exception as error:
                reason = str(error).replace("\n", " ")[:240]
                if model_index + 1 >= len(attempt_models):
                    print(json.dumps({
                        "level": "error", "jobId": job_id, "stage": "calculation-code",
                        "status": "rejected", "retryType": "generation", "reason": reason,
                    }))
                    raise
                generation_retries += 1
                print(json.dumps({
                    "level": "warning", "jobId": job_id, "stage": "calculation-code",
                    "status": "retrying", "retryType": "generation", "reason": reason,
                }))
                generation_request = (
                    request
                    + "\nThe previous code-generation response failed deterministic validation or was incomplete. "
                    + "Return a complete corrected CalculationProgram without changing task IDs. Error: "
                    + reason
                )

        if program is None:
            raise RuntimeError("CALCULATION_CODE_RETRY_EXHAUSTED")

        try:
            _emit(job_id, "calculation-execution", "started")
            output = execute_program(program.python_code, workbook_paths)
            validate_output(output, planning_output)
            _emit(job_id, "calculation-execution", "completed")
            break
        except Exception as error:
            reason = str(error).replace("\n", " ")[:240]
            if execution_corrections >= MAX_EXECUTION_CORRECTIONS:
                print(json.dumps({
                    "level": "error", "jobId": job_id, "stage": "calculation-execution",
                    "status": "rejected", "retryType": "execution-correction", "reason": reason,
                }))
                raise
            execution_corrections += 1
            print(json.dumps({
                "level": "warning", "jobId": job_id, "stage": "calculation-code",
                "status": "retrying", "retryType": "execution-correction", "reason": reason,
            }))
            if reason.startswith("CALCULATION_EXECUTION_TIMEOUT"):
                correction = (
                    "The previous valid program exceeded the 120-second execution limit. Rewrite it for bounded "
                    "linear-time execution: scan each worksheet only once, do not materialize entire worksheets "
                    "with list(...), reuse aggregates across every task, and remove repeated or nested row scans."
                )
            else:
                correction = (
                    "The previous valid program failed deterministic execution or output validation. Correct the "
                    "reported error without changing task IDs, formula IDs, metric IDs, or output fields."
                )
            request = f"{base_request}\n{correction} Error: {reason}"

    from datetime import datetime, timezone
    return CalculationArtifact(
        execution_id=str(uuid4()),
        generated_at=datetime.now(timezone.utc).isoformat(),
        duration_ms=int((time.monotonic() - started) * 1000),
        code_sha256=_sha256(program.python_code),
        python_code=program.python_code,
        tasks=output.tasks,
    )


def validate_program(source: str) -> None:
    """Reject source that can escape the tiny in-process calculation surface."""
    try:
        tree = ast.parse(source, mode="exec")
    except SyntaxError as error:
        raise ValueError("CALCULATION_CODE_SYNTAX_INVALID") from error
    forbidden_nodes = (
        ast.Global, ast.Nonlocal, ast.ClassDef, ast.AsyncFunctionDef, ast.While,
        ast.Await, ast.Yield, ast.YieldFrom, ast.AsyncWith, ast.Delete,
    )
    forbidden_names = {"__import__", "eval", "exec", "compile", "open", "input", "globals", "locals", "vars", "getattr", "setattr", "delattr", "help", "dir", "breakpoint"}
    forbidden_attributes = {"save", "read", "write", "system", "popen", "run", "spawn", "connect", "request", "urlopen"}
    function_names = [node.name for node in tree.body if isinstance(node, ast.FunctionDef)]
    if "calculate" not in function_names:
        raise ValueError("CALCULATION_CODE_MUST_DEFINE_CALCULATE")
    for node in ast.walk(tree):
        if isinstance(node, forbidden_nodes):
            raise ValueError(f"CALCULATION_CODE_UNSAFE_SYNTAX:{type(node).__name__}")
        if isinstance(node, ast.Import) and any(alias.name != "openpyxl" for alias in node.names):
            raise ValueError("CALCULATION_CODE_UNSAFE_IMPORT")
        if isinstance(node, ast.ImportFrom) and (
            node.module != "openpyxl"
            or node.level != 0
            or any(alias.name != "load_workbook" for alias in node.names)
        ):
            raise ValueError("CALCULATION_CODE_UNSAFE_IMPORT")
        if isinstance(node, ast.Name) and (node.id in forbidden_names or node.id.startswith("__")):
            raise ValueError("CALCULATION_CODE_UNSAFE_NAME")
        if isinstance(node, ast.Attribute) and (node.attr.startswith("__") or node.attr in forbidden_attributes):
            raise ValueError("CALCULATION_CODE_UNSAFE_ATTRIBUTE")
    if "load_workbook" not in source:
        raise ValueError("CALCULATION_CODE_MUST_READ_WORKBOOK")


def execute_program(source: str, workbook_paths: dict[str, Path]) -> CalculationOutput:
    context = multiprocessing.get_context("fork")
    queue: multiprocessing.Queue[dict[str, Any]] = context.Queue(maxsize=1)
    process = context.Process(target=_child_execute, args=(source, {key: str(value) for key, value in workbook_paths.items()}, queue))
    process.start()
    process.join(MAX_EXECUTION_SECONDS)
    if process.is_alive():
        process.terminate()
        process.join(5)
        raise RuntimeError("CALCULATION_EXECUTION_TIMEOUT")
    try:
        payload = queue.get(timeout=2)
    except Empty as error:
        raise RuntimeError("CALCULATION_EXECUTION_NO_OUTPUT") from error
    if payload.get("error"):
        raise RuntimeError(f"CALCULATION_EXECUTION_FAILED:{payload['error']}")
    try:
        return CalculationOutput.model_validate(payload["result"])
    except ValidationError as error:
        # ValidationError.__str__ embeds untrusted generated values. Keep logs
        # data-free while still giving the retry loop a useful correction code.
        issue = error.errors()[0] if error.errors() else {}
        raise ValueError(f"CALCULATION_OUTPUT_SCHEMA_INVALID:{issue.get('type', 'unknown')}") from error


def _child_execute(source: str, workbook_paths: dict[str, str], queue: multiprocessing.Queue[dict[str, Any]]) -> None:
    openpyxl_module = __import__("openpyxl")

    def safe_import(name: str, *_: Any, **__: Any) -> Any:
        if name == "openpyxl":
            return openpyxl_module
        raise ImportError("only openpyxl is available")

    safe_builtins = {
        "abs": abs, "all": all, "any": any, "bool": bool, "dict": dict, "enumerate": enumerate,
        "float": float, "int": int, "isinstance": isinstance, "len": len, "list": list, "max": max,
        "min": min, "range": range, "round": round, "set": set, "sorted": sorted, "str": str, "sum": sum,
        "tuple": tuple, "zip": zip, "Exception": Exception, "__import__": safe_import,
    }
    globals_scope = {"__builtins__": safe_builtins, "openpyxl": openpyxl_module}
    try:
        exec(compile(source, "<generated-calculation>", "exec"), globals_scope, globals_scope)
        result = globals_scope["calculate"](dict(workbook_paths))
        json.dumps(result, ensure_ascii=False, allow_nan=False)
        queue.put({"result": result})
    except NameError as error:
        # A missing identifier is source metadata, not workbook content. It
        # lets the corrective generation remove unsupported helpers without
        # exposing a traceback or any customer data.
        missing_name = (error.name or "unknown")[:80]
        queue.put({"error": f"NameError:{missing_name}"})
    except Exception as error:
        # Preserve only the exception class so a corrective retry can be useful
        # without placing workbook values or paths in logs.
        queue.put({"error": type(error).__name__})


def validate_output(output: CalculationOutput, planning_output: AIPlanningOutput) -> None:
    expected = {task.task_id: task for task in planning_output.calculation_plan.tasks}
    if {task.task_id for task in output.tasks} != set(expected):
        raise ValueError("CALCULATION_OUTPUT_TASK_MISMATCH")
    for result in output.tasks:
        task = expected[result.task_id]
        if result.metric_id != task.output_metric_id or result.formula_id != task.formula_id:
            raise ValueError("CALCULATION_OUTPUT_REFERENCE_MISMATCH")
        for row in result.rows:
            if not set(task.output_fields).issubset(row):
                raise ValueError("CALCULATION_OUTPUT_FIELDS_MISSING")


def _sha256(value: Any) -> str:
    data = value.encode("utf-8") if isinstance(value, str) else json.dumps(value, ensure_ascii=False, sort_keys=True).encode("utf-8")
    return hashlib.sha256(data).hexdigest()


def _emit(job_id: str | None, stage: str, status: str) -> None:
    if job_id:
        print(json.dumps({"level": "info", "jobId": job_id, "stage": stage, "status": status}))
