"""Production Fargate entry point for real Prompt and workbook planning."""

from __future__ import annotations

import argparse
import json
import os
import tempfile
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path
from typing import Any

import boto3
from botocore.config import Config
from openpyxl import load_workbook
from strands.models import BedrockModel

from .adapter import StageOutputTooLargeError, StrandsLobsterRuntimeAdapter
from .calculation import generate_and_execute
from .contracts import AIPlanningOutput

MAX_PROFILE_SHEETS = 12
MAX_PROFILE_ROWS = 4
MAX_PROFILE_COLUMNS = 30
MAX_CELL_TEXT = 240
MAX_SOURCE_REFERENCES = 160
DYNAMODB_ITEM_HARD_LIMIT_BYTES = 400 * 1024
DYNAMODB_ITEM_SAFETY_MARGIN_BYTES = 40 * 1024
MAX_STORED_JOB_ITEM_BYTES = DYNAMODB_ITEM_HARD_LIMIT_BYTES - DYNAMODB_ITEM_SAFETY_MARGIN_BYTES
# S3 downloads and openpyxl profiling are I/O-bound; a small thread pool
# parallelises multi-file workbooks without excessive memory pressure.
PARALLEL_DOWNLOAD_WORKERS = 4


class StoredPlanTooLargeError(RuntimeError):
    """Raised before a DynamoDB item can exceed its 400 KB hard limit."""

    def __init__(self, estimated_bytes: int) -> None:
        self.estimated_bytes = estimated_bytes
        super().__init__(
            f"planning job requires approximately {estimated_bytes} bytes; "
            f"safe DynamoDB budget is {MAX_STORED_JOB_ITEM_BYTES} bytes"
        )


def _dynamodb_value_size_bytes(value: Any) -> int:
    """Conservatively estimate DynamoDB's UTF-8 item accounting."""
    if value is None or isinstance(value, bool):
        return 1
    if isinstance(value, str):
        return len(value.encode("utf-8"))
    if isinstance(value, (bytes, bytearray)):
        return len(value)
    if isinstance(value, (int, float)):
        return len(str(value).encode("utf-8"))
    if isinstance(value, dict):
        return 3 + sum(
            len(str(name).encode("utf-8")) + _dynamodb_value_size_bytes(nested) + 3
            for name, nested in value.items()
        )
    if isinstance(value, (list, tuple, set)):
        return 3 + sum(_dynamodb_value_size_bytes(nested) + 3 for nested in value)
    return len(str(value).encode("utf-8"))


def candidate_item_size_bytes(
    item: dict[str, Any],
    updates: dict[str, Any],
    *,
    removals: tuple[str, ...] = (),
) -> int:
    """Estimate the complete candidate item, including names and unchanged fields."""
    candidate = {**item, **updates}
    for name in removals:
        candidate.pop(name, None)
    # Attribute names count toward the hard limit. The fixed overhead makes this
    # estimate conservative, while the separate 40 KiB margin covers encoding
    # details and future small attributes.
    return 100 + sum(
        len(name.encode("utf-8")) + _dynamodb_value_size_bytes(value) + 3
        for name, value in candidate.items()
    )


def ensure_storable_job_item(
    item: dict[str, Any],
    updates: dict[str, Any],
    *,
    removals: tuple[str, ...] = (),
) -> int:
    estimated_bytes = candidate_item_size_bytes(item, updates, removals=removals)
    if estimated_bytes > MAX_STORED_JOB_ITEM_BYTES:
        raise StoredPlanTooLargeError(estimated_bytes)
    return estimated_bytes


def profile_workbook(path: Path, upload: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    sheets: list[dict[str, Any]] = []
    references: list[dict[str, str]] = []
    try:
        for sheet in workbook.worksheets[:MAX_PROFILE_SHEETS]:
            rows: list[list[Any]] = []
            for row_index, row in enumerate(sheet.iter_rows(max_row=MAX_PROFILE_ROWS, max_col=MAX_PROFILE_COLUMNS), start=1):
                values = []
                for column_index, cell in enumerate(row, start=1):
                    value = cell.value
                    if isinstance(value, str):
                        value = value[:MAX_CELL_TEXT]
                    values.append(value)
                    if value not in (None, "") and len(references) < MAX_SOURCE_REFERENCES:
                        references.append({
                            "uploadId": upload["uploadId"], "fileName": upload["fileName"],
                            "sheetName": sheet.title, "cellRange": cell.coordinate,
                        })
                rows.append(values)
            sheets.append({
                "sheet_name": sheet.title, "max_rows_reported": sheet.max_row,
                "max_columns_reported": sheet.max_column, "sample_rows": rows,
            })
    finally:
        workbook.close()
    return {"upload_id": upload["uploadId"], "file_name": upload["fileName"], "sheets": sheets}, references


def workbook_schema(profiles: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Create an owner-visible, header-only selector catalog.

    The full workbook stays in S3 and only the signed-in owner can retrieve the
    job.  Persisting field names lets manual plan edits remain deterministic
    without sending example cell values to the browser.
    """
    catalog: list[dict[str, Any]] = []
    for profile in profiles:
        sheets = []
        for sheet in profile["sheets"]:
            header = next((row for row in sheet.get("sample_rows", []) if any(value not in (None, "") for value in row)), [])
            columns = []
            for value in header:
                if value in (None, ""):
                    continue
                name = str(value).strip()
                if name and name not in columns:
                    columns.append(name[:160])
            sheets.append({"sheetName": sheet["sheet_name"], "columns": columns[:60]})
        catalog.append({"uploadId": profile["upload_id"], "fileName": profile["file_name"], "sheets": sheets})
    return catalog


def _download_workbooks_parallel(
    s3: Any, bucket: str, uploads: list[dict[str, Any]], directory: str,
) -> dict[str, Path]:
    """Download all Excel workbooks in parallel; returns upload_id → local path."""
    paths: dict[str, Path] = {}

    def _download_one(upload: dict[str, Any]) -> tuple[str, Path]:
        target = Path(directory) / f"{upload['uploadId']}.xlsx"
        s3.download_file(bucket, upload["objectKey"], str(target))
        return upload["uploadId"], target

    with ThreadPoolExecutor(max_workers=PARALLEL_DOWNLOAD_WORKERS) as pool:
        for upload_id, path in pool.map(_download_one, uploads):
            paths[upload_id] = path
    return paths


def _download_and_profile_parallel(
    s3: Any, bucket: str, uploads: list[dict[str, Any]], directory: str,
) -> tuple[list[dict[str, Any]], list[dict[str, str]]]:
    """Download and profile all workbooks in parallel."""
    profiles: list[dict[str, Any]] = []
    references: list[dict[str, str]] = []

    def _download_and_profile(upload: dict[str, Any]) -> tuple[dict[str, Any], list[dict[str, str]]]:
        target = Path(directory) / f"{upload['uploadId']}.xlsx"
        s3.download_file(bucket, upload["objectKey"], str(target))
        return profile_workbook(target, upload)

    with ThreadPoolExecutor(max_workers=PARALLEL_DOWNLOAD_WORKERS) as pool:
        for profile, refs in pool.map(_download_and_profile, uploads):
            profiles.append(profile)
            references.extend(refs)
    return profiles, references


def run(job_id: str, operation: str) -> None:
    if operation == "CALCULATE":
        run_calculation(job_id)
        return
    table = boto3.resource("dynamodb").Table(required_env("PLANNER_TABLE"))
    s3 = boto3.client("s3", config=Config(connect_timeout=5, read_timeout=60, retries={"max_attempts": 2}))
    item = table.get_item(Key={"jobId": job_id}, ConsistentRead=True).get("Item")
    if not item:
        raise RuntimeError("JOB_NOT_FOUND")
    expected_status = "QUEUED" if operation == "CREATE" else "REVISION_QUEUED"
    table.update_item(
        Key={"jobId": job_id},
        UpdateExpression="SET #status = :running, updatedAt = :now REMOVE safeErrorCode",
        ConditionExpression="#status = :expected",
        ExpressionAttributeNames={"#status": "status"},
        ExpressionAttributeValues={":running": "RUNNING", ":expected": expected_status, ":now": utc_now()},
    )

    try:
        uploads = [upload for upload in json.loads(item["uploadManifestJson"]) if upload.get("kind", "excel") == "excel"]
        with tempfile.TemporaryDirectory(prefix="planner-") as directory:
            profiles, references = _download_and_profile_parallel(
                s3, required_env("PLANNER_INPUT_BUCKET"), uploads, directory,
            )

        request = item["prompt"]
        previous_planning_output: dict[str, Any] | None = None
        if operation == "REVISE":
            request += "\n\n使用者對上一版完整計畫的修改要求：\n" + item.get("revisionInstruction", "")
            if item.get("planningOutputJson"):
                previous_planning_output = json.loads(item["planningOutputJson"])
        base_model_id = required_env("BEDROCK_MODEL_ID")
        complex_model_id = os.environ.get("COMPLEX_BEDROCK_MODEL_ID", base_model_id)
        model = BedrockModel(
            model_id=base_model_id, region_name=required_env("AWS_REGION"),
            temperature=0.1, max_tokens=32_000,
            # Larger structured outputs need a longer streaming read window;
            # transport retries remain bounded so the stage-level retry owns recovery.
            boto_client_config=Config(connect_timeout=8, read_timeout=120, retries={"total_max_attempts": 1, "mode": "standard"}),
        )
        complex_model = BedrockModel(
            model_id=complex_model_id, region_name=required_env("AWS_REGION"),
            temperature=0.1, max_tokens=32_000,
            boto_client_config=Config(connect_timeout=8, read_timeout=120, retries={"total_max_attempts": 1, "mode": "standard"}),
        ) if complex_model_id != base_model_id else None
        plan = StrandsLobsterRuntimeAdapter(model, complex_model=complex_model).plan(
            request,
            workbook_context=profiles,
            previous_planning_output=previous_planning_output,
            job_id=job_id,
        )
        planning_output_json = plan.planning_output.model_dump_json()
        source_references_json = json.dumps(references, ensure_ascii=False)
        workbook_schema_json = json.dumps(workbook_schema(profiles), ensure_ascii=False, separators=(",", ":"))
        current_version = int(item.get("planVersion", 0))
        now = utc_now()
        item_updates = {
            "status": "NEEDS_REVIEW",
            "planVersion": current_version + 1,
            "planningOutputJson": planning_output_json,
            "sourceReferencesJson": source_references_json,
            "workbookSchemaJson": workbook_schema_json,
            "projectTitle": plan.planning_output.deck_plan.title[:300],
            "promptAlignmentScore": plan.validation_report.prompt_alignment_score,
            "updatedAt": now,
        }
        stored_bytes = ensure_storable_job_item(
            item,
            item_updates,
            removals=("revisionInstruction", "safeErrorCode"),
        )
        print(json.dumps({
            "level": "info",
            "jobId": job_id,
            "code": "DYNAMODB_ITEM_PREFLIGHT_OK",
            "estimatedBytes": stored_bytes,
            "safeLimitBytes": MAX_STORED_JOB_ITEM_BYTES,
        }))
        table.update_item(
            Key={"jobId": job_id},
            UpdateExpression="SET #status = :review, planVersion = :version, planningOutputJson = :output, sourceReferencesJson = :refs, workbookSchemaJson = :schema, projectTitle = :title, promptAlignmentScore = :score, updatedAt = :now REMOVE revisionInstruction",
            ConditionExpression="#status = :running",
            ExpressionAttributeNames={"#status": "status"},
            ExpressionAttributeValues={
                ":review": item_updates["status"], ":running": "RUNNING", ":version": item_updates["planVersion"],
                ":output": planning_output_json, ":refs": source_references_json, ":schema": workbook_schema_json,
                ":title": item_updates["projectTitle"],
                ":score": item_updates["promptAlignmentScore"],
                ":now": now,
            },
        )
    except Exception as error:
        safe_error_code = (
            "PLAN_OUTPUT_TOO_LARGE"
            if isinstance(error, StageOutputTooLargeError)
            else "PLAN_OUTPUT_STORAGE_LIMIT"
            if isinstance(error, StoredPlanTooLargeError)
            else "PLANNING_FAILED"
        )
        event: dict[str, Any] = {
            "level": "error",
            "jobId": job_id,
            "code": safe_error_code,
            "errorType": type(error).__name__,
        }
        if isinstance(error, StageOutputTooLargeError):
            event.update({
                "stage": error.stage,
                "attempts": error.attempts,
                "tokenBudget": error.token_budget,
            })
        if isinstance(error, StoredPlanTooLargeError):
            event["estimatedBytes"] = error.estimated_bytes
        print(json.dumps(event))
        table.update_item(
            Key={"jobId": job_id},
            UpdateExpression="SET #status = :failed, safeErrorCode = :code, updatedAt = :now",
            ExpressionAttributeNames={"#status": "status"},
            ExpressionAttributeValues={":failed": "FAILED", ":code": safe_error_code, ":now": utc_now()},
        )


def run_calculation(job_id: str) -> None:
    table = boto3.resource("dynamodb").Table(required_env("PLANNER_TABLE"))
    s3 = boto3.client("s3", config=Config(connect_timeout=5, read_timeout=60, retries={"max_attempts": 2}))
    item = table.get_item(Key={"jobId": job_id}, ConsistentRead=True).get("Item")
    if not item:
        raise RuntimeError("JOB_NOT_FOUND")
    attempt_id = f"{job_id}:{utc_now()}"
    table.update_item(
        Key={"jobId": job_id},
        UpdateExpression="SET #status = :running, calculationAttemptId = :attempt, updatedAt = :now REMOVE safeErrorCode",
        ConditionExpression="#status = :queued AND attribute_exists(planningOutputJson)",
        ExpressionAttributeNames={"#status": "status"},
        ExpressionAttributeValues={
            ":running": "CALCULATING",
            ":queued": "CALCULATION_QUEUED",
            ":attempt": attempt_id,
            ":now": utc_now(),
        },
    )
    try:
        planning_output = AIPlanningOutput.model_validate_json(item["planningOutputJson"])
        uploads = [upload for upload in json.loads(item["uploadManifestJson"]) if upload.get("kind", "excel") == "excel"]
        with tempfile.TemporaryDirectory(prefix="calculation-") as directory:
            workbook_paths = _download_workbooks_parallel(
                s3, required_env("PLANNER_INPUT_BUCKET"), uploads, directory,
            )
            fast_model = BedrockModel(
                model_id=required_env("BEDROCK_MODEL_ID"), region_name=required_env("AWS_REGION"),
                temperature=0.0, max_tokens=16_000,
                boto_client_config=Config(connect_timeout=8, read_timeout=45, retries={"total_max_attempts": 1, "mode": "standard"}),
            )
            fallback_model = BedrockModel(
                model_id=required_env("CALCULATION_BEDROCK_MODEL_ID"), region_name=required_env("AWS_REGION"),
                temperature=0.0, max_tokens=32_000,
                boto_client_config=Config(connect_timeout=8, read_timeout=45, retries={"total_max_attempts": 1, "mode": "standard"}),
            )
            artifact = generate_and_execute(
                fast_model, planning_output, workbook_paths,
                job_id=job_id,
                fallback_model=fallback_model,
                previous_error=item.get("calculationRetryContext"),
            )
            artifact_key = f"plans/{job_id}/calculations/{artifact.execution_id}.json"
            summary_json = json.dumps(artifact.summary(artifact_key), ensure_ascii=False, separators=(",", ":"))
            now = utc_now()
            stored_bytes = ensure_storable_job_item(
                item,
                {
                    "status": "CALCULATION_READY",
                    "calculationSummaryJson": summary_json,
                    "calculationArtifactKey": artifact_key,
                    "updatedAt": now,
                },
                removals=("safeErrorCode", "calculationRetryContext"),
            )
            print(json.dumps({
                "level": "info",
                "jobId": job_id,
                "stage": "calculation-execution",
                "code": "DYNAMODB_ITEM_PREFLIGHT_OK",
                "estimatedBytes": stored_bytes,
                "safeLimitBytes": MAX_STORED_JOB_ITEM_BYTES,
            }))
            s3.put_object(
                Bucket=required_env("PLANNER_INPUT_BUCKET"), Key=artifact_key,
                Body=json.dumps(artifact.artifact_payload(), ensure_ascii=False, separators=(",", ":")).encode("utf-8"),
                ContentType="application/json", ServerSideEncryption="AES256",
            )
        table.update_item(
            Key={"jobId": job_id},
            UpdateExpression="SET #status = :ready, calculationSummaryJson = :summary, calculationArtifactKey = :artifact, updatedAt = :now REMOVE safeErrorCode, calculationRetryContext, calculationAttemptId",
            ConditionExpression="#status = :running AND calculationAttemptId = :attempt",
            ExpressionAttributeNames={"#status": "status"},
            ExpressionAttributeValues={
                ":ready": "CALCULATION_READY",
                ":running": "CALCULATING",
                ":attempt": attempt_id,
                ":summary": summary_json,
                ":artifact": artifact_key,
                ":now": now,
            },
        )
    except Exception as error:
        error_reason = str(error)
        code = (
            "PLAN_OUTPUT_STORAGE_LIMIT"
            if isinstance(error, StoredPlanTooLargeError)
            else "CALCULATION_EXECUTION_TIMEOUT"
            if error_reason.startswith("CALCULATION_EXECUTION_TIMEOUT")
            else "CALCULATION_CODE_REJECTED"
            if error_reason.startswith("CALCULATION_CODE_")
            else "CALCULATION_FAILED"
        )
        retry_context = error_reason[:240] if error_reason.startswith("CALCULATION_") else code
        event = {
            "level": "error",
            "jobId": job_id,
            "stage": "calculation-execution",
            "code": code,
            "errorType": type(error).__name__,
        }
        if isinstance(error, StoredPlanTooLargeError):
            event["estimatedBytes"] = error.estimated_bytes
        print(json.dumps(event))
        table.update_item(
            Key={"jobId": job_id},
            UpdateExpression="SET #status = :failed, safeErrorCode = :code, calculationRetryContext = :retryContext, updatedAt = :now REMOVE calculationAttemptId",
            ConditionExpression="#status = :running AND calculationAttemptId = :attempt",
            ExpressionAttributeNames={"#status": "status"},
            ExpressionAttributeValues={
                ":failed": "CALCULATION_FAILED",
                ":running": "CALCULATING",
                ":attempt": attempt_id,
                ":code": code,
                ":retryContext": retry_context,
                ":now": utc_now(),
            },
        )


def utc_now() -> str:
    from datetime import datetime, timezone
    return datetime.now(timezone.utc).isoformat()


def required_env(name: str) -> str:
    value = os.environ.get(name)
    if not value:
        raise RuntimeError(f"MISSING_{name}")
    return value


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--job-id", default=os.environ.get("PLANNER_JOB_ID"))
    parser.add_argument("--operation", choices=["CREATE", "REVISE", "CALCULATE"], default=os.environ.get("PLANNER_OPERATION"))
    arguments = parser.parse_args()
    if not arguments.job_id or not arguments.operation:
        parser.error("job id and operation are required")
    run(arguments.job_id, arguments.operation)


if __name__ == "__main__":
    main()
