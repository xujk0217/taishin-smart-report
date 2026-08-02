"""Deterministic data profiling and evidence construction for arbitrary Excel files."""

from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .presentation_contracts import ChartData, ChartSeries, ClaimRecord, EvidenceMetric, EvidencePacketV2, TableData
from .universal_pipeline_contracts import ColumnProfile, SheetProfile


MAX_ROWS = 500
MAX_EXAMPLE_VALUES = 3


def profile_excel_files(paths: list[str | Path]) -> list[SheetProfile]:
    profiles: list[SheetProfile] = []
    for workbook_path in paths:
        path = Path(workbook_path)
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        try:
            for sheet in workbook.worksheets:
                rows = list(sheet.iter_rows(max_row=MAX_ROWS, values_only=True))
                if not rows:
                    continue
                header_index = _find_header_row(rows)
                headers = _headers(rows[header_index])
                data_rows = rows[header_index + 1 :]
                profiles.append(SheetProfile(
                    file_name=path.name,
                    sheet_name=sheet.title,
                    header_row=header_index + 1,
                    row_count=len([row for row in data_rows if any(cell not in (None, "") for cell in row)]),
                    columns=_profile_columns(headers, data_rows),
                    quality_findings=_quality_findings(headers, data_rows),
                ))
        finally:
            workbook.close()
    return profiles


def build_evidence_from_profiles(profiles: list[SheetProfile]) -> EvidencePacketV2:
    metrics: list[EvidenceMetric] = []
    charts: list[ChartData] = []
    tables: list[TableData] = []
    claims: list[ClaimRecord] = []
    for sheet_index, profile in enumerate(profiles, start=1):
        numeric_columns = [column for column in profile.columns if column.inferred_type == "number" and column.numeric_sum is not None]
        text_columns = [column for column in profile.columns if column.inferred_type == "text"]
        display_numeric_columns = numeric_columns[:4]
        metric_ids_for_sheet: list[str] = []
        for column in display_numeric_columns:
            metric_id = _safe_id(f"metric_{sheet_index}_{column.name}")
            metric_ids_for_sheet.append(metric_id)
            metrics.append(EvidenceMetric(
                metric_id=metric_id,
                label=f"{profile.sheet_name} {column.name} 合計",
                value=float(column.numeric_sum or 0),
                unit="",
                display_format="number",
                source_refs=[f"{profile.file_name}#{profile.sheet_name}!{column.name}"],
                calculation=f"sum numeric values in column {column.name}",
            ))
        if numeric_columns:
            primary_numeric = numeric_columns[0]
            chart_id = _safe_id(f"chart_{sheet_index}_{primary_numeric.name}")
            categories = [column.name for column in display_numeric_columns]
            values = [float(column.numeric_sum or 0) for column in display_numeric_columns]
            charts.append(ChartData(
                chart_id=chart_id,
                title=f"{profile.sheet_name} 數值欄位合計",
                chart_type="column",
                categories=categories,
                series=[ChartSeries(name="合計", values=values)],
                metric_refs=metric_ids_for_sheet,
            ))
            table_id = _safe_id(f"table_{sheet_index}_{primary_numeric.name}")
            tables.append(TableData(
                table_id=table_id,
                title=f"{profile.sheet_name} 數值摘要",
                headers=["欄位", "合計"],
                rows=[[name, value] for name, value in zip(categories, values)],
                source_refs=[f"{profile.file_name}#{profile.sheet_name}"],
            ))
            if metrics:
                claims.append(ClaimRecord(
                    claim_id=_safe_id(f"claim_{sheet_index}_summary"),
                    text=f"{{{{{metric_ids_for_sheet[0]}}}}} 是此資料表中可追溯的主要數值摘要之一。",
                    metric_refs=[metric_ids_for_sheet[0]],
                    chart_refs=[chart_id],
                    source_refs=[f"{profile.file_name}#{profile.sheet_name}"],
                ))
        elif text_columns:
            tables.append(TableData(
                table_id=_safe_id(f"table_{sheet_index}_text_profile"),
                title=f"{profile.sheet_name} 文字欄位摘要",
                headers=["欄位", "樣本"],
                rows=[[column.name, "、".join(column.examples)] for column in text_columns[:5]],
                source_refs=[f"{profile.file_name}#{profile.sheet_name}"],
            ))
    return EvidencePacketV2(
        packet_id="evp-" + hashlib.sha256("|".join(f"{p.file_name}:{p.sheet_name}:{p.row_count}" for p in profiles).encode("utf-8")).hexdigest()[:16],
        metrics=metrics,
        charts=charts,
        claims=claims,
        tables=tables,
    )


def _find_header_row(rows: list[tuple[Any, ...]]) -> int:
    best_index = 0
    best_score = -1
    for index, row in enumerate(rows[:20]):
        score = sum(1 for cell in row if isinstance(cell, str) and cell.strip())
        if score > best_score:
            best_index = index
            best_score = score
    return best_index


def _headers(row: tuple[Any, ...]) -> list[str]:
    headers: list[str] = []
    for index, value in enumerate(row):
        text = str(value).strip() if value not in (None, "") else f"Column{index + 1}"
        headers.append(text[:80])
    return headers


def _profile_columns(headers: list[str], rows: list[tuple[Any, ...]]) -> list[ColumnProfile]:
    columns: list[ColumnProfile] = []
    for index, header in enumerate(headers):
        values = [row[index] for row in rows if index < len(row) and row[index] not in (None, "")]
        examples = [str(value)[:80] for value in values[:MAX_EXAMPLE_VALUES]]
        numeric_values = [_as_number(value) for value in values]
        numeric_values = [value for value in numeric_values if value is not None]
        inferred = "empty"
        numeric_sum = None
        if values and len(numeric_values) / len(values) >= 0.8:
            inferred = "number"
            numeric_sum = float(sum(numeric_values))
        elif values:
            inferred = "text"
        columns.append(ColumnProfile(
            name=header,
            inferred_type=inferred,
            non_empty_count=len(values),
            examples=examples,
            numeric_sum=numeric_sum,
        ))
    return columns


def _quality_findings(headers: list[str], rows: list[tuple[Any, ...]]) -> list[str]:
    findings: list[str] = []
    if len(headers) != len(set(headers)):
        findings.append("duplicate column names detected")
    if not rows:
        findings.append("no data rows detected")
    return findings


def _as_number(value: Any) -> float | None:
    if isinstance(value, bool):
        return None
    if isinstance(value, int | float):
        return float(value)
    if isinstance(value, str):
        cleaned = value.replace(",", "").replace("%", "").replace("％", "").strip()
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _safe_id(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() else "_" for ch in value.lower())
    while "__" in cleaned:
        cleaned = cleaned.replace("__", "_")
    return cleaned.strip("_")[:80] or "item"
